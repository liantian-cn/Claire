#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# __author__ = 'Lian Tian'
# __email__ = "liantian.me+code@gmail.com"
#
# GNU GENERAL PUBLIC LICENSE

import pathlib
import logging
import time
from copy import copy
import smtplib
import logging
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email.utils import formataddr
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email import encoders
from openpyxl import load_workbook
import urllib.parse
import datetime

BASE_PATH = pathlib.Path(__file__).parent
CONFIG_PATH = BASE_PATH.joinpath("配置文件")
CONFIG_FILE = CONFIG_PATH.joinpath("当前配置.xlsx")
TEMPLATE_FILE = CONFIG_PATH.joinpath("当前模板.xlsx")
INPUT_PATH = BASE_PATH.joinpath("输入目录")
OUTPUT_PATH = BASE_PATH.joinpath("输出目录")
TEMP_OUTPUT_PATH = OUTPUT_PATH.joinpath(str(int(time.time())))


def load_config():
    workbook = load_workbook(CONFIG_FILE)
    error = False
    _base_config = dict()
    config_sheet = workbook["基础信息表"]
    for row_number in range(2, config_sheet.max_row + 1):
        conf_attribute = config_sheet.cell(row=row_number, column=1).value
        conf_value = config_sheet.cell(row=row_number, column=2).value
        required = config_sheet.cell(row=row_number, column=3).value == "是"
        if (conf_value is None) and (required is True):
            print(f"{conf_attribute} 值是必要的，请检查配置文件")
            error = True
        if conf_attribute is None:
            break
        if conf_value is None:
            _base_config[conf_attribute.upper().strip()] = conf_value
        else:
            _base_config[conf_attribute.upper().strip()] = conf_value.strip()

    if _base_config["MAIL_PROXY"] is None:
        _base_config["MAIL_PROXY"] = _base_config["MAIL_SENDER"]
    else:
        print("代表邮件发送已启用")

    if _base_config["MAIL_SENDER_DISPLAY"] is None:
        _base_config["msg_From"] = _base_config["MAIL_PROXY"]
    else:
        _base_config["msg_From"] = formataddr((_base_config["MAIL_SENDER_DISPLAY"], _base_config["MAIL_PROXY"]))

    if (_base_config["MAIL_CC"] is not None) and (_base_config["MAIL_CC_DISPLAY"] is not None):
        _base_config["msg_CC"] = formataddr((_base_config["MAIL_CC_DISPLAY"], _base_config["MAIL_CC"]))
    elif (_base_config["MAIL_CC"] is not None) and (_base_config["MAIL_CC_DISPLAY"] is None):
        _base_config["msg_CC"] = formataddr((_base_config["MAIL_CC"], _base_config["MAIL_CC"]))
    else:
        _base_config["msg_CC"] = None

    if (_base_config["MAIL_BCC"] is not None) and (_base_config["MAIL_BCC_DISPLAY"] is not None):
        _base_config["msg_BCC"] = formataddr((_base_config["MAIL_BCC_DISPLAY"], _base_config["MAIL_BCC"]))
    elif (_base_config["MAIL_BCC"] is not None) and (_base_config["MAIL_BCC_DISPLAY"] is None):
        _base_config["msg_BCC"] = formataddr((_base_config["msg_BCC"], _base_config["MAIL_BCC"]))
    else:
        _base_config["msg_BCC"] = None

    _base_config["MAIL_SERVER_PORT"] = int(_base_config["MAIL_SERVER_PORT"])
    _base_config["SOURCE_HEADER_LINE_NO"] = int(_base_config["SOURCE_HEADER_LINE_NO"])
    _base_config["TEMPLATE_HEADER_LINE_NO"] = int(_base_config["TEMPLATE_HEADER_LINE_NO"])

    if _base_config["DEBUG_MODE"].lower().strip() == "true":
        _base_config["DEBUG_MODE"] = True
    else:
        _base_config["DEBUG_MODE"] = False

    if error:
        print("基础配置文件有误，程序将要退出")
        time.sleep(10)
        exit(0)

    _address_book = dict()
    address_book_sheet = workbook["邮件及收件人信息对照表"]

    for row_number in range(2, address_book_sheet.max_row):
        title = address_book_sheet.cell(row=row_number, column=1).value
        address = address_book_sheet.cell(row=row_number, column=2).value
        display_name = address_book_sheet.cell(row=row_number, column=3).value
        if (title is None) or (address is None):
            break
        if display_name is None:
            display_name = address
        if _address_book.get(title, None) is None:
            _address_book[title] = {"msg_To": [formataddr((display_name, address))], "receivers": [address]}
        else:
            _address_book[title]["msg_To"].append(formataddr((display_name, address)))
            _address_book[title]["receivers"].append(address)

    workbook.close()

    _wanted_keys = []
    workbook = load_workbook(TEMPLATE_FILE)
    worksheet = workbook.worksheets[0]
    for column_number in range(1, worksheet.max_column + 1):
        key = worksheet.cell(row=_base_config["TEMPLATE_HEADER_LINE_NO"], column=column_number).value
        if key is None:
            break
        _wanted_keys.append(key)
    workbook.close()

    return _base_config, _address_book, _wanted_keys


def read_source_file(_header_line, _focus):
    workbook = load_workbook(INPUT_PATH.joinpath("当前输入.xlsx"))
    worksheet = workbook.worksheets[0]

    keys = [worksheet.cell(1, col_index).value for col_index in range(1, worksheet.max_column + 1)]

    _source_list = []

    for row_index in range(_header_line + 1, worksheet.max_row + 1):
        d = {keys[col_index - 1]: worksheet.cell(row_index, col_index).value for col_index in
             range(1, worksheet.max_column + 1)}
        _source_list.append(d)

    _focus_set = set(line[_focus] for line in _source_list)
    if None in _focus_set:
        _focus_set.remove(None)

    return _source_list, _focus_set


def copy_style(new_cell, old_cell):
    if old_cell.has_style:
        new_cell.font = copy(old_cell.font)
        new_cell.border = copy(old_cell.border)
        new_cell.fill = copy(old_cell.fill)
        new_cell.number_format = copy(old_cell.number_format)
        new_cell.protection = copy(old_cell.protection)
        new_cell.alignment = copy(old_cell.alignment)


def gen_split_file(_base_config, _wanted_keys, _source_list, _focus_set):
    _result = []
    TEMP_OUTPUT_PATH.mkdir(exist_ok=True)
    for _focus in _focus_set:
        filter_list = [x for x in _source_list if x[_base_config["SOURCE_HEADER_FOCUS"]] == _focus]
        output_name = "".join([c for c in _focus if c.isalpha() or c.isdigit() or c == ' ']).strip()
        output_file = TEMP_OUTPUT_PATH.joinpath("{}.xlsx".format(output_name))
        workbook = load_workbook(TEMPLATE_FILE)
        worksheet = workbook.worksheets[0]
        for line in range(0, len(filter_list)):
            for col_num in range(0, len(_wanted_keys)):
                new_cell = worksheet.cell(row=line + _base_config["TEMPLATE_HEADER_LINE_NO"] + 1, column=col_num + 1,
                                          value=filter_list[line][_wanted_keys[col_num]])
                old_cell = worksheet.cell(row=_base_config["TEMPLATE_HEADER_LINE_NO"] + 1, column=col_num + 1)

                copy_style(new_cell, old_cell)

        workbook.save(output_file)
        workbook.close()
        _result.append((_focus, output_file))
    return _result


def main():
    base_config, address_book, wanted_keys = load_config()

    # print(base_config)
    # print(address_book)
    # print(wanted_keys)

    source_list, focus_set = read_source_file(_header_line=base_config["SOURCE_HEADER_LINE_NO"],
                                              _focus=base_config["SOURCE_HEADER_FOCUS"])
    # print(source_list)
    # print(focus_set)
    TEMP_OUTPUT_PATH.mkdir(exist_ok=True)
    need_to_send_files = gen_split_file(base_config, wanted_keys, source_list, focus_set)


if __name__ == '__main__':
    main()

    # for focus in focus_set:
    #     filter_list = [x for x in source_list if x[base_config["SOURCE_HEADER_FOCUS"]] == focus]
    #     output_name = "".join([c for c in focus if c.isalpha() or c.isdigit() or c == ' ']).strip()
    #     output_file = TEMP_OUTPUT_PATH.joinpath("{}.xlsx".format(output_name))
    #     # print(filter_list)
